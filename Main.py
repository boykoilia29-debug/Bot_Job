import telebot
from telebot import types
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

# Получаем токен из переменных окружения
TOKEN = os.environ.get('TOKEN')
if not TOKEN:
    # Для локального тестирования можно указать токен напрямую
    TOKEN = '8712850028:AAFHOX1kcw_5vNYryKjp0KxR3G8YmhJigtg'  # Вставьте ваш токен

bot = telebot.TeleBot(TOKEN)

# Состояния пользователей
user_states = {}
user_data = {}

# Множество для хранения ID пользователей, уже отправивших заявку
# ВАЖНО: При перезапуске бота это множество очистится!
# Для продакшена лучше хранить в Excel/БД
submitted_users = set()

# Функция для инициализации Excel файла
def init_excel():
    if not os.path.exists('applications.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"
        headers = ['Дата', 'User ID', 'Username', 'Желаемая должность', 'Опыт работы', 'Статус']
        ws.append(headers)
        wb.save('applications.xlsx')
        
        # Создаем отдельный лист для отслеживания пользователей
        ws_users = wb.create_sheet("Пользователи")
        ws_users.append(['User ID', 'Username', 'Дата заявки', 'Статус'])
        wb.save('applications.xlsx')

# Функция для проверки, отправлял ли пользователь заявку
def has_user_submitted(user_id):
    # Проверяем в памяти
    if user_id in submitted_users:
        return True
    
    # Проверяем в Excel файле
    try:
        wb = openpyxl.load_workbook('applications.xlsx')
        if 'Пользователи' in wb.sheetnames:
            ws_users = wb['Пользователи']
            for row in range(2, ws_users.max_row + 1):
                if ws_users.cell(row=row, column=1).value == user_id:
                    submitted_users.add(user_id)  # Кэшируем в память
                    return True
    except:
        pass
    
    return False

# Функция для отметки пользователя как отправившего заявку
def mark_user_as_submitted(user_id, username):
    submitted_users.add(user_id)
    
    try:
        wb = openpyxl.load_workbook('applications.xlsx')
        if 'Пользователи' not in wb.sheetnames:
            ws_users = wb.create_sheet("Пользователи")
            ws_users.append(['User ID', 'Username', 'Дата заявки', 'Статус'])
        else:
            ws_users = wb['Пользователи']
        
        next_row = ws_users.max_row + 1
        ws_users.cell(row=next_row, column=1, value=user_id)
        ws_users.cell(row=next_row, column=2, value=username)
        ws_users.cell(row=next_row, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws_users.cell(row=next_row, column=4, value='Заявка отправлена')
        
        wb.save('applications.xlsx')
    except Exception as e:
        print(f"Ошибка при сохранении пользователя: {e}")

# Функция для сохранения данных в Excel
def save_to_excel(user_id, username, position, experience):
    try:
        wb = openpyxl.load_workbook('applications.xlsx')
        ws = wb.active
        
        next_row = ws.max_row + 1
        
        ws.cell(row=next_row, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=next_row, column=2, value=user_id)
        ws.cell(row=next_row, column=3, value=username)
        ws.cell(row=next_row, column=4, value=position)
        ws.cell(row=next_row, column=5, value=experience)
        ws.cell(row=next_row, column=6, value='Новая')
        
        wb.save('applications.xlsx')
        
        # Отмечаем пользователя как отправившего заявку
        mark_user_as_submitted(user_id, username)
        
        return True
    except Exception as e:
        print(f"Ошибка при сохранении заявки: {e}")
        return False

# Функция для отправки финального сообщения с благодарностью и ссылками
def send_thank_you_message(chat_id):
    # ЗДЕСЬ ВСТАВЬТЕ ВАШИ ССЫЛКИ (3 штуки)
    links = """
🔗 <b>Полезные ссылки:</b>

1. <a href="https://t.me/YPOLN0MOCHEN">Наш Telegram канал</a> - будьте в курсе новостей
2. <a href="https://t.me/+1lSqpBU0vtUyNTZh">Наша команда(обязательно подать заявку)</a>
3. <a href="https://t.me/CTPAX_D0K7EP0B">Мой ЛС</a> - если есть вопросы
    """
    
    thank_you_text = f"""
🎉 <b>Спасибо за вашу заявку!</b>

Мы получили вашу анкету и уже начали её обработку. 
Наши специалисты свяжутся с вами в ближайшее время (обычно в течение 24 часов).

Пока вы ждёте ответа, предлагаем ознакомиться с нашими ресурсами:

{links}

💬 Если у вас возникнут вопросы, не стесняйтесь обращаться!
    """
    
    # Создаем клавиатуру с кнопками для ссылок
    markup = types.InlineKeyboardMarkup(row_width=1)
    
    # ЗДЕСЬ ВСТАВЬТЕ ВАШИ ССЫЛКИ В КНОПКИ
    btn1 = types.InlineKeyboardButton("📱 Наш Telegram канал", url="https://t.me/YPOLN0MOCHEN")
    btn2 = types.InlineKeyboardButton("🌐 Наша команда(обязательно подать заявку)", url="https://t.me/+1lSqpBU0vtUyNTZh")
    btn3 = types.InlineKeyboardButton("💬 Мой ЛС", url="https://t.me/CTPAX_D0K7EP0B")
    
    markup.add(btn1, btn2, btn3)
    
    bot.send_message(chat_id, thank_you_text, parse_mode='HTML', reply_markup=markup)

# Приветственное сообщение
@bot.message_handler(commands=['start'])
def send_welcome(message):
    user_id = message.chat.id
    
    # Проверяем, отправлял ли пользователь уже заявку
    if has_user_submitted(user_id):
        # Если уже отправлял, показываем сообщение с благодарностью и ссылками
        send_thank_you_message(user_id)
        return
    
    # ЗДЕСЬ НУЖНО ЗАПОЛНИТЬ СПИСОК ВАКАНСИЙ
    vacancies_list = """
    • Трейдер
    • Закупщик НФТ или крипты
    • Программист
    • Арбитражник
    • Трафер
    • СММ
    """
    
    welcome_text = f"""
🎯 <b>Добро пожаловать в бот для приема заявок!</b>

📋 <b>Актуальные вакансии:</b>
{vacancies_list}

⚠️ <b>Важно:</b> Вы можете отправить только ОДНУ заявку. 
Пожалуйста, заполните анкету внимательно.

Для подачи заявки нажмите кнопку ниже 👇
    """
    
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn = types.KeyboardButton("📝 Подать заявку")
    markup.add(btn)
    
    bot.send_message(message.chat.id, welcome_text, parse_mode='HTML', reply_markup=markup)

# Начало процесса подачи заявки
@bot.message_handler(func=lambda message: message.text == "📝 Подать заявку")
def start_application(message):
    user_id = message.chat.id
    
    # Проверяем, не отправлял ли пользователь уже заявку
    if has_user_submitted(user_id):
        send_thank_you_message(user_id)
        return
    
    user_states[user_id] = 'awaiting_username'
    
    bot.send_message(
        user_id, 
        "Шаг 1 из 3\n\n"
        "Пожалуйста, отправьте ваш Telegram username (например: @username):",
        reply_markup=types.ReplyKeyboardRemove()
    )

# Обработчик для получения username
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'awaiting_username')
def get_username(message):
    user_id = message.chat.id
    
    # Дополнительная проверка
    if has_user_submitted(user_id):
        del user_states[user_id]
        send_thank_you_message(user_id)
        return
    
    username = message.text.strip()
    
    # Проверка формата username
    if not username.startswith('@'):
        username = '@' + username
    
    user_data[user_id] = {'username': username}
    user_states[user_id] = 'awaiting_position'
    
    bot.send_message(user_id, "Шаг 2 из 3\n\nКем вы хотите работать?")

# Обработчик для получения желаемой должности
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'awaiting_position')
def get_position(message):
    user_id = message.chat.id
    
    # Дополнительная проверка
    if has_user_submitted(user_id):
        del user_states[user_id]
        send_thank_you_message(user_id)
        return
    
    position = message.text.strip()
    
    user_data[user_id]['position'] = position
    user_states[user_id] = 'awaiting_experience'
    
    bot.send_message(user_id, "Шаг 3 из 3\n\nЕсть ли у вас опыт работы? (Если да, напишите сколько)")

# Обработчик для получения опыта и сохранения заявки
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == 'awaiting_experience')
def get_experience(message):
    user_id = message.chat.id
    
    # Финальная проверка
    if has_user_submitted(user_id):
        del user_states[user_id]
        send_thank_you_message(user_id)
        return
    
    experience = message.text.strip()
    
    # Сохраняем данные
    user_data[user_id]['experience'] = experience
    
    # Сохраняем в Excel
    success = save_to_excel(
        user_id,
        user_data[user_id]['username'],
        user_data[user_id]['position'],
        experience
    )
    
    # Очищаем состояния
    if user_id in user_states:
        del user_states[user_id]
    if user_id in user_data:
        del user_data[user_id]
    
    if success:
        # Отправляем финальное сообщение с благодарностью и ссылками
        send_thank_you_message(user_id)
    else:
        bot.send_message(
            user_id,
            "❌ Произошла ошибка при сохранении заявки. Пожалуйста, попробуйте позже или свяжитесь с поддержкой."
        )

# Команда для администратора - просмотр статистики
@bot.message_handler(commands=['chakApplication'])
def admin_panel(message):
    admin_id = 8347600681  # ЗАМЕНИТЕ НА ВАШ ID
    
    if message.chat.id == admin_id:
        try:
            wb = openpyxl.load_workbook('applications.xlsx')
            ws = wb.active
            
            total_applications = ws.max_row - 1
            unique_users = len(submitted_users)
            
            stats_text = f"""
📊 <b>Статистика бота:</b>

👥 Всего пользователей: {unique_users}
📝 Всего заявок: {total_applications}
            """
            
            bot.send_message(message.chat.id, stats_text, parse_mode='HTML')
            
            if ws.max_row > 1:
                response = "📋 <b>Последние 5 заявок:</b>\n\n"
                
                start_row = max(2, ws.max_row - 4)
                for row in range(start_row, ws.max_row + 1):
                    date = ws.cell(row=row, column=1).value
                    username = ws.cell(row=row, column=3).value
                    position = ws.cell(row=row, column=4).value
                    
                    response += f"📅 {date}\n"
                    response += f"👤 {username}\n"
                    response += f"💼 {position}\n"
                    response += "─" * 20 + "\n"
                
                bot.send_message(message.chat.id, response, parse_mode='HTML')
        except Exception as e:
            bot.send_message(message.chat.id, f"Ошибка при чтении статистики: {e}")
    else:
        bot.send_message(message.chat.id, "У вас нет доступа к этой команде")

# Команда для администратора - выгрузить базу
@bot.message_handler(commands=['getbase'])
def get_database(message):
    admin_id = 8347600681  # ЗАМЕНИТЕ НА ВАШ ID
    
    if message.chat.id == admin_id:
        try:
            with open('applications.xlsx', 'rb') as file:
                bot.send_document(message.chat.id, file, caption="📊 База заявок")
        except:
            bot.send_message(message.chat.id, "Файл с заявками не найден")
    else:
        bot.send_message(message.chat.id, "У вас нет доступа к этой команде")

# Инициализация Excel при запуске
init_excel()

# Запуск бота
if __name__ == '__main__':
    print("Бот запущен...")
    print(f"Отслеживаем {len(submitted_users)} пользователей в памяти")
    bot.infinity_polling()

